# MOISES TORO
# PROPER AI


#Importing libraries
import xlwings as xw
from openpyxl.utils import get_column_letter

#Reading schedules
user = input("Enter your laptop user: ")
mec = input("Enter Current Closing Month (Example for March:'03'):")


fund1 = xw.Book(r'C:\Users\file 1'.format(user=user, folder=mec,file=mec))
fund2 = xw.Book(r'C:\Users\file 2'.format(user=user, folder=mec,file=mec))
fund3 = xw.Book(r'C:\Users\file 3'.format(user=user, folder=mec,file=mec))
fund4 = xw.Book(r'C:\Users\file 4'.format(user=user, folder=mec,file=mec))
fund5 = xw.Book(r'C:\Users\file 5'.format(user=user, folder=mec,file=mec))
fund6 = xw.Book(r'C:\Users\file 6'.format(user=user, folder=mec,file=mec))
fund7 = xw.Book(r'C:\Users\file 7'.format(user=user, folder=mec,file=mec))
fund8 = xw.Book(r'C:\Users\file 8'.format(user=user, folder=mec,file=mec))
fund9 = xw.Book(r'C:\Users\file 9'.format(user=user, folder=mec,file=mec))

files_to_update = [fund1, fund2, fund3, fund4, fund5, fund6, fund7, fund8, fund9]





#Reading Yardi report
source_file = xw.Book(r'source_file.xlsx')
source = source_file.sheets.active  # in specific book


#Updating every schedule in files_to_update
cell_h = source['h1']
cell_i = source['i1']



for i in range(len(files_to_update)):

    current_schedule = files_to_update[i]
    
    sheet_names = [sheet.name for sheet in current_schedule.sheets]

    #Updating date for all properties

    for a_cell in sheet_names:
        if  a_cell != 'Revision':
            current_schedule.sheets['{:}'.format(a_cell)]['B4'].value = source_file['D8'].value


    #Updating schedule keys        
    schedule_keys = []

    for a_cell in source_file["A6"].expand("down"):
        
        if a_cell.value in sheet_names and (cell_i.offset(a_cell.row-1).value != 0 or cell_h.offset(a_cell.row-1).value != 0):
            for i in range(1,current_schedule.sheets[a_cell.value].range('b' + str(current_schedule.sheets[a_cell.value].cells.last_cell.row)).end('up').row):
                if current_schedule.sheets[a_cell.value]['B{}'.format(i)].value == '2023 Beginning Balance':
                    first_cell_date = i
                    break

            for i in range(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date)].end('down').row -  current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date)].row + 1):
                
                row_range = current_schedule.sheets[a_cell.value]['A7:bb7']
                row_values = row_range.value


                #Amount index
                def amount_index():
                    global amounts_index
                    if current_schedule.name == '2023-{file} - fund1.xlsx'.format(file=mec):
                        amounts_index = row_values.index("1000")
                    elif current_schedule.name == '2023-{file} - fund2.xlsx'.format(file=mec):
                        amounts_index = row_values.index("1001")
                    elif current_schedule.name == '2023-{file} - fund3.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1002')
                    elif current_schedule.name == '2023-{file} - fund4.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1003')
                    elif current_schedule.name == '2023-{file} - fund5.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1004')
                    elif current_schedule.name == '2023-{file} - fund6.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1005')
                    elif current_schedule.name == '2023-{file} - fund7.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1006')
                    elif current_schedule.name == '2023-{file} - fund8.xlsx'.format(file=mec):
                        #amounts_index = row_values.index('SF Multifamily IV JV LLC(2011)')
                        amounts_index = row_values.index('1007')
                    elif current_schedule.name == '2023-{file} - fund8.xlsx'.format(file=mec):
                        amounts_index = row_values.index('1008')
                
                amount_index()

                                
                amounts_cell = row_range[0, amounts_index]
                total_index = row_values.index("Total Running Balance")
                total_cell = row_range[0, total_index]
                des_index = row_values.index("Description")
                des_cell = row_range[0, des_index]
                #regex = -round(len(str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value)) * 0.6)

                

                row_range = current_schedule.sheets[a_cell.value].range(f"{get_column_letter(amounts_cell.column)}{first_cell_date  + i}:{get_column_letter(total_cell.column - 1)}{first_cell_date + i}")
                #amount = sum(value for value in row_range.value if value is not None)
                
                for amt in row_range:
                    if amt.value is not None:
                        #schedule_keys.append(str(a_cell.value)+str(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date + i)].value) + str(amt.value))
                        #schedule_keys.append(str(a_cell.value)+str(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date + i)].value) + str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value)[-round(len(str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value)) * 0.6):] + str(amt.value))
                        #schedule_keys.append(str(a_cell.value)+str(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date + i)].value) +  str(amt.value))
                        schedule_keys.append(str(a_cell.value)+str(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date + i)].value) + str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value) + str(amt.value))



                
                
                #schedule_keys.append(str(a_cell.value)+str(current_schedule.sheets[a_cell.value]['b{}'.format(first_cell_date + i)].value) + str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value)[-round(len(str(des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(first_cell_date + i))].value)) * 0.6):] + str(amount))


    #Updating schedules
    for a_cell in source_file["A6"].expand("down"):
        if a_cell.value in sheet_names and cell_h.offset(a_cell.row-1).value != 0:
                    

            row_range = current_schedule.sheets[a_cell.value]['A7:bb7']
            row_values = row_range.value

            amount_index()

            date_index = row_values.index("Post Month")
            des_index = row_values.index("Description")
            total_index = row_values.index("Total Running Balance")
            date_cell = row_range[0, date_index]
            des_cell = row_range[0, des_index]
            total_cell = row_range[0, total_index]


            if not str(a_cell.value)+str(source_file['d1'](a_cell.row).value) + str(source_file['k1'](a_cell.row).value) + str(cell_h(a_cell.row).value ) in schedule_keys:

                
                amounts_cell = row_range[0, amounts_index]

                current_schedule.sheets[a_cell.value].range('{}:{}'.format(date_cell.end('down').row+1, date_cell.end('down').row+1)).insert('down') #insert a row

                amounts_cell.sheet[amounts_cell.address.replace(str(amounts_cell.row), str(date_cell.end('down').offset(row_offset=1).row))].value = cell_h.offset(a_cell.row-1).value
                
                date_cell.end('down').offset(row_offset=1).value = source_file['D1'].offset(a_cell.row-1).value
                
                des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(date_cell.end('down').row))].value = source_file['K1'].offset(a_cell.row-1).value
                
                total_cell.sheet[total_cell.address.replace(str(total_cell.row), str(date_cell.end('down').row))].value = total_cell.sheet[total_cell.address.replace(str(total_cell.row), str(date_cell.end('down').row-1))].value + amounts_cell.sheet[amounts_cell.address.replace(str(amounts_cell.row), str(date_cell.end('down').row))].value



        if a_cell.value in sheet_names and cell_i.offset(a_cell.row-1).value != 0:


            row_range = current_schedule.sheets[a_cell.value]['A7:bb7']
            row_values = row_range.value


            amount_index()

            date_index = row_values.index("Post Month")
            des_index = row_values.index("Description")
            total_index = row_values.index("Total Running Balance")
            date_cell = row_range[0, date_index]
            des_cell = row_range[0, des_index]
            total_cell = row_range[0, total_index]


            if not str(a_cell.value)+str(source_file['d1'](a_cell.row).value)  + str(source_file['k1'](a_cell.row).value) +  str(cell_i(a_cell.row).value *-1) in schedule_keys:

                
                amounts_cell = row_range[0, amounts_index]
                current_schedule.sheets[a_cell.value].range('{}:{}'.format(date_cell.end('down').row+1, date_cell.end('down').row+1)).insert('down') #insert a row

                amounts_cell.sheet[amounts_cell.address.replace(str(amounts_cell.row), str(date_cell.end('down').offset(row_offset=1).row))].value = cell_i.offset(a_cell.row-1).value * -1

                date_cell.end('down').offset(row_offset=1).value = source_file['D1'].offset(a_cell.row-1).value

                des_cell.sheet[des_cell.address.replace(str(des_cell.row), str(date_cell.end('down').row))].value = source_file['K1'].offset(a_cell.row-1).value

                total_cell.sheet[total_cell.address.replace(str(total_cell.row), str(date_cell.end('down').row))].value = total_cell.sheet[total_cell.address.replace(str(total_cell.row), str(date_cell.end('down').row-1))].value + amounts_cell.sheet[amounts_cell.address.replace(str(amounts_cell.row), str(date_cell.end('down').row))].value

    print(str(current_schedule),'updated')

